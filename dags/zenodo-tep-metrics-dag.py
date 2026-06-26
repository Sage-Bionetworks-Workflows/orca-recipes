"""Zenodo TREAT-AD Target Enabling Package (TEP) Metrics DAG

This DAG automates the monthly collection of TEP usage metrics (views and
downloads) from the Zenodo API for the TREAT-AD community and exports them to
Synapse. The DAG follows these steps:

1. Fetch target enabling/enablement records from the Zenodo API using an
   authentication token stored in Airflow Variables (paginated, with Airflow
   retries handling transient/busy API responses).
2. Validate the pulled metrics against an expected schema to ensure stability of
   the Zenodo API response and the computed aggregates. If the schema has
   changed or a value is missing, the task logs the problem and fails so that
   Airflow retries kick in and the failure alert is sent.
3. Export the metrics to a styled Excel workbook report and upload it to a Synapse
   folder.
4. Notify the collaborator(s) via email that a new report is available.

On any task failure (after retries are exhausted), a Synapse message is sent to
the DPE/developer list via the DAG's on_failure_callback so that schema
changes or other breakages are surfaced.

The DAG runs monthly. The Zenodo API token is stored in the Airflow secrets 
backend as the ZENODO_API_TOKEN Variable.
"""

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Tuple

from airflow.decorators import dag, task
from airflow.models import Param, Variable
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import requests
from synapseclient import File

from src.synapse_hook import SynapseHook

logger = logging.getLogger(__name__)

# Zenodo records API endpoint
ZENODO_RECORDS_URL = "https://zenodo.org/api/records"

# Title search terms used to discover TEP records in Zenodo. Both spellings are
# searched because records use either "enabling" or "enablement".
SEARCH_TERMS = ['"target enabling"', '"target enablement"']

# Zenodo community that TEP records must belong to
TREATAD_COMMUNITY_ID = "treatad"

# Fields we expect on every processed record. Used by the validation task to
# detect Zenodo API schema drift
REQUIRED_RECORD_FIELDS = [
    "title",
    "date",
    "views",
    "unique_views",
    "downloads",
    "unique_downloads",
    "link",
    "category",
]

# Numeric metric fields that must be non-negative integers
METRIC_FIELDS = ["views", "unique_views", "downloads", "unique_downloads"]

# Headers for the Excel report (in the order they appear in the workbook)
EXCEL_HEADERS = [
    "Title",
    "Publication Date",
    "Total Views",
    "Unique Views",
    "Total Downloads",
    "Unique Downloads",
    "Link",
]

dag_params = {
    "synapse_conn_id": Param("SYNAPSE_ORCA_SERVICE_ACCOUNT_CONN", type="string"),
    # Synapse folder/project the Excel report is uploaded to. Update to the
    # shared TREAT-AD reporting location.
    "synapse_export_folder": Param("syn75951837", type="string"),
    # Comma-separated Synapse usernames of collaborator(s) to notify when the
    # monthly report is available. (No support for python lists in Params.)
    "collaborator_user_list": Param("3460442", type="string"),
    # Comma-separated Synapse usernames or IDs of DPE/developers to 
    # alert on failure.
    "dev_user_list": Param("3460442", type="string"),
}

dag_config = {
    # Run on the first day of the month at midnight
    "schedule_interval": "0 0 1 * *", 
    "start_date": datetime(2025, 1, 1),
    "catchup": False,
    "default_args": {
        "retries": 3,
    },
    "tags": ["zenodo", "treat-ad"],
    "params": dag_params,
}


def fetch_tep_records(
    api_token: str,
    search_terms: List[str] = SEARCH_TERMS,
    community_id: str = TREATAD_COMMUNITY_ID,
) -> List[Dict[str, Any]]:
    """Fetch and process TREAT-AD target enabling/enablement records from Zenodo.

    Searches the Zenodo records API for each search term (handling pagination),
    de-duplicates by record id, filters to the given community, and truncates each
    record down to the metric fields of interest. Each record is tagged with a
    category of "report" (Package/Report) or "component"
    (Component/Resource) based on its title.

    Arguments:
        api_token (str): Zenodo API token
        search_terms (List[str]): Title search terms to query
        community_id (str): Zenodo community id that records must belong to

    Returns:
        List[Dict[str, Any]]: Processed TEP records

    Raises:
        requests.exceptions.RequestException: If a Zenodo API request fails, this
            is intentionally allowed to propagate so Airflow retries handle a
            busy/transient API.
    """
    headers = {"Authorization": f"Bearer {api_token}"}
    all_records: List[Dict[str, Any]] = []

    for search_term in search_terms:
        logger.info(f"Searching Zenodo for {search_term}")
        params: Dict[str, Any] = {
            "q": f"metadata.title:{search_term}",
            "size": 100,
            "page": 1,
        }
        while True:
            logger.info(f"  Fetching page {params['page']} for {search_term}")
            response = requests.get(
                ZENODO_RECORDS_URL, params=params, headers=headers, timeout=30
            )
            response.raise_for_status()
            data = response.json()

            hits = data["hits"]["hits"]
            if not hits:
                break

            all_records.extend(hits)
            logger.info(f"    Retrieved {len(hits)} records")

            total_records = data["hits"]["total"]
            if len(all_records) >= total_records or len(hits) < params["size"]:
                break
            params["page"] += 1

    logger.info(f"Retrieved {len(all_records)} total records (pre-dedup)")

    # Remove duplicates (same record id)
    seen_ids = set()
    unique_records = []
    for record in all_records:
        if record["id"] not in seen_ids:
            seen_ids.add(record["id"])
            unique_records.append(record)
    logger.info(f"After removing duplicates: {len(unique_records)} unique records")

    teps: List[Dict[str, Any]] = []
    for record in unique_records:
        communities = record.get("metadata", {}).get("communities", [])
        if not any(comm.get("id") == community_id for comm in communities):
            continue  # Skip non-TREAT-AD records

        title = record["metadata"]["title"]
        stats = record.get("stats", {})
        teps.append(
            {
                "title": title,
                "date": record["metadata"]["publication_date"],
                "views": stats.get("views", 0),
                "unique_views": stats.get("unique_views", 0),
                "downloads": stats.get("downloads", 0),
                "unique_downloads": stats.get("unique_downloads", 0),
                "link": f"https://zenodo.org/records/{record['id']}",
                "category": categorize_title(title),
            }
        )

    logger.info(f"Found {len(teps)} TREAT-AD TEP records after community filtering")
    return teps


def categorize_title(title: str) -> str:
    """Classify a record as a TEP report or a supporting component by title.

    Records whose title mentions "component" or "resource" are supporting
    materials, everything else (including "package"/"report") is treated as a
    main TEP report.

    Arguments:
        title (str): The record title

    Returns:
        str: "component" for supporting materials, otherwise "report"
    """
    title_lower = title.lower()
    if "component" in title_lower or "resource" in title_lower:
        return "component"
    return "report"


def validate_records(records: List[Dict[str, Any]]) -> None:
    """Validate the pulled TEP metrics for schema stability.

    Logs all validation issues and only raises at the end if at least one check
    fails, so the Airflow logs contain maximum debug info while still failing the
    run (and triggering the failure alert) on bad output.

    Checks performed:
      1) At least one record was returned.
      2) Every record contains all REQUIRED_RECORD_FIELDS
      3) Metric fields are non-negative integers
      4) title, date and link fields have non-empty values

    Arguments:
        records (List[Dict[str, Any]]): Processed records from fetch_tep_records

    Raises:
        ValueError: If one or more validation checks fail with the details of 
            all failures
    """
    errors: List[str] = []

    if not records:
        msg = "Validation failed: Zenodo returned zero TREAT-AD TEP records."
        logger.error(msg)
        errors.append(msg)

    for i, record in enumerate(records):
        missing = [f for f in REQUIRED_RECORD_FIELDS if f not in record]
        if missing:
            msg = f"Validation failed: record {i} missing fields: {missing}"
            logger.error(msg)
            errors.append(msg)

        for field in METRIC_FIELDS:
            value = record.get(field)
            if not isinstance(value, int) or isinstance(value, bool) or value < 0:
                msg = (
                    f"Validation failed: record {i} field '{field}' is not a "
                    f"non-negative integer (got {value!r})."
                )
                logger.error(msg)
                errors.append(msg)

        for field in ["title", "date", "link"]:
            if not record.get(field):
                msg = f"Validation failed: record {i} has empty '{field}'."
                logger.error(msg)
                errors.append(msg)

    if errors:
        raise ValueError(
            "Zenodo TEP metrics validation failed:\n- " + "\n- ".join(errors)
        )

    logger.info(f"Validation passed for {len(records)} records.")


def build_workbook(records: List[Dict[str, Any]], filepath: str) -> Tuple[int, int]:
    """Build a Excel workbook of TEP metrics and save it locally.

    Creates two worksheets ("TEP Reports" and "TEP Components") split by record
    category, each with a header row, rows of data, and a totals row.

    Arguments:
        records (List[Dict[str, Any]]): Validated TEP records
        filepath (str): Path to write the .xlsx file to

    Returns:
        Dict[str, int]: Counts of report records and component records
    """
    main_teps = [r for r in records if r["category"] == "report"]
    supporting = [r for r in records if r["category"] == "component"]

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    totals_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    totals_font = Font(bold=True)
    column_widths = {"A": 60, "B": 18, "C": 12, "D": 12, "E": 15, "F": 17, "G": 50}

    for sheet_name, sheet_records in [
        ("TEP Reports", main_teps),
        ("TEP Components", supporting),
    ]:
        ws = wb.create_sheet(title=sheet_name)

        ws.append(EXCEL_HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in sheet_records:
            ws.append(
                [
                    record["title"],
                    record["date"],
                    record["views"],
                    record["unique_views"],
                    record["downloads"],
                    record["unique_downloads"],
                    record["link"],
                ]
            )

        ws.append([])  # spacer row
        totals_row_idx = len(sheet_records) + 3
        ws.append(
            [
                "TOTALS",
                "",
                sum(r["views"] for r in sheet_records),
                sum(r["unique_views"] for r in sheet_records),
                sum(r["downloads"] for r in sheet_records),
                sum(r["unique_downloads"] for r in sheet_records),
                "",
            ]
        )
        for cell in ws[totals_row_idx]:
            cell.fill = totals_fill
            cell.font = totals_font

        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

    wb.save(filepath)
    logger.info(
        f"Wrote workbook to {filepath} ({len(main_teps)} reports, {len(supporting)} components)"
    )
    return {
        "report_records": len(main_teps),
        "component_records": len(supporting),
    }


def _send_synapse_message(
    conn_id: str, usernames: List[str], subject: str, body: str
) -> None:
    """Send a Synapse message (delivered as email) to a list of usernames.

    Arguments:
        conn_id (str): Synapse connection id
        usernames (List[str]): Synapse usernames to message
        subject (str): Message subject
        body (str): Message body
    """
    usernames = [u.strip() for u in usernames if u.strip()]
    if not usernames:
        logger.warning(f"No Synapse usernames provided; skipping message: {subject}")
        return
    client = SynapseHook(conn_id).client
    owner_ids = [client.getUserProfile(u).get("ownerId") for u in usernames]
    client.sendMessage(owner_ids, subject, body)
    logger.info(f"Sent Synapse message '{subject}' to {usernames}")


def alert_on_failure(context: Dict[str, Any]) -> None:
    """DAG on_failure_callback that alerts DPE/developers via Synapse message
    through email when a task fails after retries are exhausted.

    Arguments:
        context (Dict[str, Any]): Airflow task context for the failed task
    """
    params = context.get("params", {})
    dev_user_list = params.get("dev_user_list", "")
    task_instance = context.get("task_instance")
    exception = context.get("exception")
    task_id = getattr(task_instance, "task_id", "unknown_task")
    log_url = getattr(task_instance, "log_url", "")
    dag_id = getattr(task_instance, "dag_id", getattr(context.get("dag"), "dag_id", "unknown_dag"))
    run_id = getattr(task_instance, "run_id", context.get("run_id", "unknown_run"))
    # logical_date is the modern key; fall back to execution_date for older runs.
    execution_date = context.get("logical_date") or context.get("execution_date") or ""

    subject = f"Zenodo TEP Metrics DAG failure: {dag_id}"
    body = (
        f"The Zenodo TEP metrics DAG '{dag_id}' failed on task '{task_id}'.\n\n"
        f"Run ID: {run_id}\n"
        f"Execution date: {execution_date}\n\n"
        f"Exception: {exception}\n\n"
        f"Logs: {log_url}\n\n"
        "This may indicate a Zenodo API schema change or a transient API issue. "
        "Please review the task logs."
    )
    try:
        _send_synapse_message(
            conn_id=params.get("synapse_conn_id", ""), usernames=dev_user_list.split(","), subject=subject, body=body
        )
    except Exception:
        logger.exception("Failed to send Synapse failure alert.")


@dag(on_failure_callback=alert_on_failure, **dag_config)
def zenodo_tep_metrics_dag():
    """Pull TREAT-AD TEP metrics from Zenodo, validate, export to Synapse, notify."""

    @task
    def fetch_metrics(**context) -> List[Dict[str, Any]]:
        """Fetch TREAT-AD TEP records from the Zenodo API.

        Returns:
            List[Dict[str, Any]]: Processed TEP records
        """
        api_token = Variable.get("ZENODO_API_TOKEN")
        return fetch_tep_records(api_token)

    @task
    def validate_metrics(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Validate the pulled metrics for schema stability.

        Arguments:
            records (List[Dict[str, Any]]): Records from fetch_metrics

        Returns:
            List[Dict[str, Any]]: The validated records
        """
        validate_records(records)
        return records

    @task
    def export_to_synapse(records: List[Dict[str, Any]], **context) -> Dict[str, str]:
        """Build the Excel report and upload it to Synapse.

        Arguments:
            records (List[Dict[str, Any]]): Validated TEP records
            **context: Airflow task context containing DAG parameters

        Returns:
            Dict[str, str]: Uploaded file info (id and name)
        """
        run_date = context["ds_nodash"]  # YYYYMMDD of the logical run date
        filename = f"TREATAD_Target_Enabling_Metrics_{run_date}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)

        build_workbook(records, filepath)
        try:
            hook = SynapseHook(context["params"]["synapse_conn_id"])
            uploaded = hook.client.store(
                File(filepath, parent=context["params"]["synapse_export_folder"])
            )
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)

        logger.info(f"Uploaded {filename} to Synapse as {uploaded.id}")
        return {"id": uploaded.id, "name": filename}

    @task
    def notify_collaborators(file_info: Dict[str, str], **context) -> None:
        """Notify collaborator(s) that the monthly report is available.

        Arguments:
            file_info (Dict[str, str]): Uploaded file info from export_to_synapse.
            **context: Airflow task context containing DAG parameters
        """
        subject = "New TREAT-AD Target Enabling Metrics report available"
        body = (
            f"A new monthly TREAT-AD target enabling metrics report "
            f"('{file_info['name']}') has been uploaded to Synapse: "
            f"https://www.synapse.org/Synapse:{file_info['id']}"
        )
        _send_synapse_message(
            conn_id=context["params"]["synapse_conn_id"],
            usernames=context["params"]["collaborator_user_list"].split(","),
            subject=subject,
            body=body,
        )

    records = fetch_metrics()
    validated = validate_metrics(records)
    file_info = export_to_synapse(validated)
    notify_collaborators(file_info)


zenodo_tep_metrics_dag()